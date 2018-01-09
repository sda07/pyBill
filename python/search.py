import openpyxl as op
from openpyxl.utils import coordinate_from_string, column_index_from_string
from calculate import calculate_it, income_tax, savings
import msvcrt
# ...

wb_1 = op.load_workbook('savings.xlsx')
wb = op.load_workbook('DATA13.xlsx')

def find_header(sheet_name, wb):
	''' to find the coloumn name which 
	is uniformly stated in row 1 of every
	sheet 
	'''
	#wb = op.load_workbook(filename = 'DATA13.xlsx')

	sheet = wb.get_sheet_by_name(sheet_name)
	header = list(sheet.rows)[0]
	hd_dic = {}

	#print(header)
	for counter in header:
		hd_dic[counter.value] =column_index_from_string(counter.column)

	#print(hd_dic)
	return hd_dic

def find_name(sheet_name, emp_name, wb):
	''' find name and return row number
	'''
	#wb = op.load_workbook(filename = 'DATA13.xlsx')
	sheet = wb.get_sheet_by_name(sheet_name)
	row_val = 0

	for rw in sheet.iter_rows():
		for cl in rw:
		    if cl.value == emp_name.upper():
		    	row_val = cl.row
	return row_val

def show_current_savings(sheet_name, emp_name):
	''' show in console current data '''
	sheet = wb_1.get_sheet_by_name(sheet_name)
	header = find_header(sheet_name, wb_1)
	name_row = find_name(sheet_name, emp_name, wb_1)
	savings_dict ={}

	for temp in list(header.keys()):
		print(temp + ' : ' + str(sheet.cell(row = name_row, 
			column = header[temp]).value) or 0 + '\n')
		savings_dict[temp] = sheet.cell(row = name_row, 
			column = header[temp]).value
	return savings_dict


def find_net(sheet_name, row_num, wb):
	''' find basic, basic1(Gp), splpay, qpay,ta, cca(da_on_ta)
	hra, da, other1(WA), itax, sc, cghs, grinsurance, gpf_t(nps)
	'''
	net_dict = {}
	header = find_header(sheet_name, wb)
	
	#wb = op.load_workbook(filename = 'DATA13.xlsx')

	sheet = wb.get_sheet_by_name(sheet_name)
	try:
		net_dict['BASIC'] = sheet.cell(row = row_num, 
		column = header['BASIC']).value
		net_dict['BASIC1'] =  sheet.cell(row = row_num, 
		column = header['BASIC1']).value
		net_dict['SPLPAY'] =  sheet.cell(row = row_num, 
		column = header['SPLPAY']).value
		net_dict['QPAY'] =  sheet.cell(row = row_num, 
		column = header['QPAY']).value
		net_dict['TA'] =  sheet.cell(row = row_num, 
		column = header['TA']).value
		net_dict['CCA'] =  sheet.cell(row = row_num, 
		column = header['CCA']).value
		net_dict['HRA'] =  sheet.cell(row = row_num, 
		column = header['HRA']).value
		net_dict['DA'] =  sheet.cell(row = row_num, 
		column = header['DA']).value
		net_dict['WA'] =  sheet.cell(row = row_num, 
		column = header['WA']).value
		net_dict['OTHER1'] =  sheet.cell(row = row_num, 
		column = header['OTHER1']).value
		net_dict['ITAX'] =  sheet.cell(row = row_num, 
		column = header['ITAX']).value
		net_dict['SC'] =  sheet.cell(row = row_num, 
		column = header['SC']).value
		net_dict['CGHS'] =  sheet.cell(row = row_num, 
		column = header['CGHS']).value
		net_dict['GRINSURANC'] =  sheet.cell(row = row_num, 
		column = header['GRINSURANC']).value
		net_dict['GPFT'] =  sheet.cell(row = row_num, 
		column = header['GPFT']).value
		net_dict['LIC'] =  sheet.cell(row = row_num, 
		column = header['LIC']).value
		net_dict['ROP'] =  sheet.cell(row = row_num, 
		column = header['ROP']).value
		net_dict['PTAX'] =  sheet.cell(row = row_num, 
		column = header['PTAX']).value
		net_dict['CGHS'] = sheet.cell(row = row_num,
			column = header['CGHS']).value

		
	except ValueError:
		pass # for value not found in the worksheet

		
	return net_dict







#print (wb.sheetnames)
def main():
	str_in = input("Enter value:")
	#wb = op.load_workbook(filename = 'DATA13.xlsx')
	total = {}
	for sh in wb.sheetnames:
		#print(sh)

		row_num = find_name(sh, str_in, wb)
		all_details = find_net(sh, row_num, wb)
		
		
		
		for val in list(all_details.keys()):

			
			try:
				total[val] += all_details[val]

			except KeyError:
				total = all_details
				break
			except TypeError:
				pass
		# print(total['ROP'])
		#print(total)
	#sheet = wb_1.get_sheet_by_name('Sheet1')
	sav = show_current_savings('Sheet1', str_in)
	# print(all_details)
	# print(total)
	# msvcrt.getch()
	summary = calculate_it(total)
	cal_sav = savings(sav, summary)
	print(summary)
	income_tax(summary, cal_sav)

if __name__ == "__main__":
	main() 


	#print(find_name(sh,str_in))
	# for rw in wb.get_sheet_by_name(sh).iter_rows():
	# 	for cl in rw:
	# 		if cl.value == str_in.upper():
	# 			print (sh + ' ' + cl.column + str(cl.row)





