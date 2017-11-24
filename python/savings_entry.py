import openpyxl as op
from openpyxl.utils import coordinate_from_string, column_index_from_string
from search import find_header, find_name
wb = op.load_workbook('savings.xlsx')

def show_current_savings(sheet_name, emp_name, wb):
	''' show in console current data '''
	sheet = wb.get_sheet_by_name(sheet_name)
	header = find_header(sheet_name, wb)
	name_row = find_name(sheet_name, emp_name, wb)
	savings_dict ={}

	for temp in list(header.keys()):
		print(temp + ' : ' + str(sheet.cell(row = name_row, 
			column = header[temp]).value) or 0 + '\n')
		savings_dict[temp] = sheet.cell(row = name_row, 
			column = header[temp]).value
	return savings_dict

def data_entry(sheet_name, emp_name):
	sheet = wb.get_sheet_by_name(sheet_name)
	header = find_header(sheet_name, wb)
	name_row = find_name(sheet_name, emp_name, wb)

	for temp in list(header.keys()):
		if temp not in ['NAME'] :
			
			try:
				str_in = input( '\n' + temp + ' :')
			except SyntaxError:
				str_in = None

			print (str_in)

			if (str_in is None) or (str_in.strip() == ''):
				print(sheet.cell(row = name_row, 
					column = header[temp]).value)
			else:
				print('\n\n')
				try:
					sheet.cell(row = name_row, 
						column = header[temp]).value = int(str_in)
				except ValueError:
					print('Enter valid input')
					break



	wb.save(filename = 'savings.xlsx')







str_in = input('Enter Name : ')
show_current_savings('Sheet1',str_in, wb)
data_entry('Sheet1',str_in)
